"""
match_employers.py
-------------------
Matches employers between two sources:

  File A - "unique" employer list   (data/raw/unique_employers.xlsx)
            columns: employerid, employername, _freq

  File B - "lightcast" panel data   (data/raw/lightcast_employers.xlsx)
            columns: employer_id, emp_name, year, freq
            (multiple rows per employer: repeated across years 2015-2019,
             plus duplicate / inconsistently-named entries)

Pipeline
--------
1. Clean employer names in both files (lowercase, strip punctuation and
   common legal suffixes such as Inc/LLC/Corp) so "Walmart" and
   "Walmart Inc." collapse to the same comparable string.
2. Collapse File B's duplicates: for every (employer_id, cleaned name),
   sum freq across duplicate rows, then keep only the single YEAR with
   the highest freq for that employer_id.
3. Where several employer_ids in File B reduce to the *same* cleaned
   name (naming inconsistency), keep only the employer_id/name with the
   highest freq - that becomes File B's one candidate row per company.
4. Fuzzy-match every File A employer against the cleaned File B
   candidates (token-sort ratio), keeping the single best-scoring
   candidate per File A employer.
5. Flag valid_match = TRUE when the similarity score clears the
   threshold (default 90). Rows below the threshold are still written
   out (valid_match = FALSE) so they can be spot-checked, matching how
   this kind of matching review is normally QA'd.

Run:
    python scripts/match_employers.py
Output:
    data/output/matched_employers.xlsx
"""

import re
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SIMILARITY_THRESHOLD = 90  # matches below this score are flagged invalid

LEGAL_SUFFIXES = [
    r"\bincorporated\b", r"\binc\b", r"\bllc\b", r"\bl l c\b", r"\bltd\b",
    r"\blimited\b", r"\bcorp\b", r"\bcorporation\b", r"\bco\b", r"\bcompany\b",
    r"\bgroup\b", r"\bholdings\b", r"\bplc\b", r"\bpllc\b", r"\blp\b", r"\bllp\b",
]


def clean_name(name: str) -> str:
    """Lowercase, strip punctuation, drop common legal suffixes, collapse whitespace."""
    if not isinstance(name, str):
        return ""
    text = name.lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^\w\s']", " ", text)  # strip punctuation except apostrophes
    for suffix in LEGAL_SUFFIXES:
        text = re.sub(suffix, " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_file_a(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    df["cleaned_company"] = df["employername"].apply(clean_name)
    return df


def load_and_collapse_file_b(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    df["cleaned_match"] = df["emp_name"].apply(clean_name)

    # Step 2: sum duplicate (employer_id, name, year) rows, then keep only
    # the highest-freq year per employer_id.
    grouped = (
        df.groupby(["employer_id", "emp_name", "cleaned_match", "year"], as_index=False)
        ["freq"].sum()
    )
    idx_best_year = grouped.groupby("employer_id")["freq"].idxmax()
    best_year_per_id = grouped.loc[idx_best_year].reset_index(drop=True)

    # Step 3: several employer_ids can share a cleaned_match (inconsistent
    # naming) - keep only the highest-freq employer_id per cleaned name.
    idx_best_name = best_year_per_id.groupby("cleaned_match")["freq"].idxmax()
    collapsed = best_year_per_id.loc[idx_best_name].reset_index(drop=True)
    return collapsed


def match(df_a: pd.DataFrame, df_b_collapsed: pd.DataFrame) -> pd.DataFrame:
    choices = df_b_collapsed["cleaned_match"].tolist()

    results = []
    for _, row_a in df_a.iterrows():
        query = row_a["cleaned_company"]
        best = process.extractOne(query, choices, scorer=fuzz.token_sort_ratio)

        if best is None:
            match_row = None
            score = 0
        else:
            matched_text, score, match_pos = best
            match_row = df_b_collapsed.iloc[match_pos]

        results.append({
            "employer_name (file A unique)": row_a["employername"],
            "employer_id (file A unique)": row_a["employerid"],
            "freq (file A unique)": row_a["_freq"],
            "employer_name (lightcast)": match_row["emp_name"] if match_row is not None else None,
            "employer_id (lightcast)": match_row["employer_id"] if match_row is not None else None,
            "freq (lightcast)": match_row["freq"] if match_row is not None else None,
            "year": match_row["year"] if match_row is not None else None,
            "similarity_score": round(score, 1),
            "cleaned_company": query,
            "cleaned_match": match_row["cleaned_match"] if match_row is not None else None,
            "valid_match": bool(score >= SIMILARITY_THRESHOLD),
        })

    return pd.DataFrame(results)


def format_output(path: str) -> None:
    """Light formatting pass: bold header, autofilter, frozen header row,
    readable column widths, and a green/amber fill on valid_match."""
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    body_font = Font(name="Arial")
    true_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    false_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = [cell.value for cell in ws[1]]
    valid_match_col = headers.index("valid_match") + 1

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
        match_cell = row[valid_match_col - 1]
        match_cell.fill = true_fill if match_cell.value else false_fill

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(14, min(34, len(str(header)) + 4))
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


if __name__ == "__main__":
    df_a = load_file_a("data/raw/unique_employers.xlsx")
    df_b_collapsed = load_and_collapse_file_b("data/raw/lightcast_employers.xlsx")

    output = match(df_a, df_b_collapsed)
    output = output.sort_values(
        ["valid_match", "freq (file A unique)"], ascending=[False, False]
    ).reset_index(drop=True)

    output.to_excel("data/output/matched_employers.xlsx", index=False)
    format_output("data/output/matched_employers.xlsx")

    n_valid = output["valid_match"].sum()
    print(f"Matched {n_valid} / {len(output)} File A employers "
          f"at a similarity threshold of {SIMILARITY_THRESHOLD}.")
    print(output.head(10).to_string(index=False))
