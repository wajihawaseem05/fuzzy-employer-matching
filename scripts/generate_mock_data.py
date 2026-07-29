"""
generate_mock_data.py
----------------------
Creates two SYNTHETIC Excel files that stand in for the original,
confidential project inputs:

  File A -> data/raw/unique_employers.xlsx
            (mirrors the "unique data" file: one row per employer,
             with a total frequency count)

  File B -> data/raw/lightcast_employers.xlsx
            (mirrors the "lightcast data" file: multiple rows per
             employer across years 2015-2019, including duplicate /
             inconsistently-named entries)

None of the names, IDs, or counts below are real. They exist only to
demonstrate the matching pipeline in match_employers.py. Row counts are
intentionally scaled down (100s, not 82,000+ / millions) for a fast,
reviewable GitHub demo — the pipeline logic scales to the original size.
"""

import random
import numpy as np
import pandas as pd

random.seed(42)
np.random.seed(42)

# ---------------------------------------------------------------------
# 1. A small "universe" of fictional companies, each with a canonical
#    name plus a few realistic name VARIANTS (legal suffixes, typos,
#    abbreviations) - this is what makes fuzzy matching necessary.
# ---------------------------------------------------------------------
COMPANIES = [
    {"canonical": "Bluepeak Health Systems",
     "variants_a": ["Bluepeak Health Systems, Inc."],
     "variants_b": ["Bluepeak Health Systems", "Bluepeak Health", "Bluepeak Healthsystems"]},
    {"canonical": "Norwell Staffing Group",
     "variants_a": ["Norwell Staffing Group LLC"],
     "variants_b": ["Norwell Staffing Group", "Norwell Staffing", "Norwell Staffing Grp"]},
    {"canonical": "Castlebrook Financial",
     "variants_a": ["Castlebrook Financial, Inc."],
     "variants_b": ["Castlebrook Financial", "Castlebrook Finl", "Castlebrook Financial Co"]},
    {"canonical": "Prairie & Vine Foods",
     "variants_a": ["Prairie & Vine Foods Corp."],
     "variants_b": ["Prairie & Vine Foods", "Prairie and Vine Foods", "Prairie Vine Foods"]},
    {"canonical": "Ashford Technologies",
     "variants_a": ["Ashford Technologies LLC"],
     "variants_b": ["Ashford Technologies", "Ashford Tech", "Ashford Technolgies"]},
    {"canonical": "Kestrel Logistics",
     "variants_a": ["Kestrel Logistics, Inc."],
     "variants_b": ["Kestrel Logistics", "Kestrel Logisitcs", "Kestrel Logistic Services"]},
    {"canonical": "Harborview Insurance",
     "variants_a": ["Harborview Insurance Group"],
     "variants_b": ["Harborview Insurance", "Harborview Ins", "Harborview Insurance Grp"]},
    {"canonical": "Milltown Manufacturing",
     "variants_a": ["Milltown Manufacturing Co."],
     "variants_b": ["Milltown Manufacturing", "Milltown Mfg", "Milltown Manufacturing Company"]},
    {"canonical": "Solace Behavioral Care",
     "variants_a": ["Solace Behavioral Care"],
     "variants_b": ["Solace Behavioral Care", "Solace Behavioral", "Solace Behavioural Care"]},
    {"canonical": "Ridgeline Consulting Partners",
     "variants_a": ["Ridgeline Consulting Partners, LLC"],
     "variants_b": ["Ridgeline Consulting Partners", "Ridgeline Consulting", "Ridgeline Consulting Prtnrs"]},
    {"canonical": "Fenwick Retail Holdings",
     "variants_a": ["Fenwick Retail Holdings Inc."],
     "variants_b": ["Fenwick Retail Holdings", "Fenwick Retail", "Fenwick Retail Hldgs"]},
    {"canonical": "Copper Creek Energy",
     "variants_a": ["Copper Creek Energy Corp."],
     "variants_b": ["Copper Creek Energy", "Copper Creek Engy", "Copper Creek Energy Co"]},
    {"canonical": "Northbridge Legal Services",
     "variants_a": ["Northbridge Legal Services PLLC"],
     "variants_b": ["Northbridge Legal Services", "Northbridge Legal", "Northbridge Lgl Svcs"]},
    {"canonical": "Alder & Finch Advertising",
     "variants_a": ["Alder & Finch Advertising, Inc."],
     "variants_b": ["Alder & Finch Advertising", "Alder and Finch Advertising", "Alder Finch Advertising"]},
    {"canonical": "Summit Peak Airlines",
     "variants_a": ["Summit Peak Airlines"],
     "variants_b": ["Summit Peak Airlines", "Summit Peak Air", "Summitpeak Airlines"]},
    {"canonical": "Glenmoor Senior Living",
     "variants_a": ["Glenmoor Senior Living LLC"],
     "variants_b": ["Glenmoor Senior Living", "Glenmoor Senior Care", "Glenmoor Snr Living"]},
    {"canonical": "Ironvale Steel Works",
     "variants_a": ["Ironvale Steel Works, Inc."],
     "variants_b": ["Ironvale Steel Works", "Ironvale Steelworks", "Ironvale Steel"]},
    {"canonical": "Brightfield Education Group",
     "variants_a": ["Brightfield Education Group"],
     "variants_b": ["Brightfield Education Group", "Brightfield Education", "Brightfield Edu Grp"]},
    {"canonical": "Wexley Pharmaceuticals",
     "variants_a": ["Wexley Pharmaceuticals, Inc."],
     "variants_b": ["Wexley Pharmaceuticals", "Wexley Pharma", "Wexley Pharmaceutical"]},
    {"canonical": "Dunmore Hospitality Group",
     "variants_a": ["Dunmore Hospitality Group LLC"],
     "variants_b": ["Dunmore Hospitality Group", "Dunmore Hospitality", "Dunmore Hosp Group"]},
]

# A handful of employers that exist ONLY in File A (no match expected) and
# ONLY in File B (no match expected) - real projects always have these.
A_ONLY = ["Thistledown Farms Cooperative", "Marrow & Oak Design Studio", "Quillfeather Media, LLC"]
B_ONLY = ["Yellowstone Grain Traders", "Petrel Marine Services", "Auburn Hill Bakery"]

YEARS = [2015, 2016, 2017, 2018, 2019]


def build_file_a():
    rows = []
    next_id = 1000
    for c in COMPANIES:
        name = random.choice(c["variants_a"])
        rows.append({
            "employerid": next_id,
            "employername": name,
            "_freq": np.random.randint(20, 4000),
        })
        next_id += 7
    for name in A_ONLY:
        rows.append({
            "employerid": next_id,
            "employername": name,
            "_freq": np.random.randint(5, 150),
        })
        next_id += 7
    df = pd.DataFrame(rows).sort_values("_freq", ascending=False).reset_index(drop=True)
    return df


def build_file_b():
    rows = []
    next_id = 5000
    for c in COMPANIES:
        # Each real-world company shows up under 1-3 name variants in
        # Lightcast, each with its OWN employer_id, and each variant
        # repeats across several years with different (duplicate) counts.
        n_variants = random.choice([1, 1, 2, 2, 3])
        chosen_variants = random.sample(c["variants_b"], k=min(n_variants, len(c["variants_b"])))
        for variant_name in chosen_variants:
            emp_id = next_id
            next_id += 3
            years_present = sorted(random.sample(YEARS, k=random.randint(2, 5)))
            for yr in years_present:
                rows.append({
                    "employer_id": emp_id,
                    "emp_name": variant_name,
                    "year": yr,
                    "freq": np.random.randint(10, 5000),
                })
                # occasional exact duplicate row for the same id/name/year
                # (mirrors the "duplicates because of the same rows" note)
                if random.random() < 0.15:
                    rows.append({
                        "employer_id": emp_id,
                        "emp_name": variant_name,
                        "year": yr,
                        "freq": np.random.randint(10, 5000),
                    })
    for name in B_ONLY:
        emp_id = next_id
        next_id += 3
        for yr in random.sample(YEARS, k=random.randint(2, 4)):
            rows.append({
                "employer_id": emp_id,
                "emp_name": name,
                "year": yr,
                "freq": np.random.randint(10, 2000),
            })
    df = pd.DataFrame(rows).sample(frac=1, random_state=42).reset_index(drop=True)
    return df


def _format_raw(path: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header_len = len(str(col_cells[0].value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, header_len + 4)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


if __name__ == "__main__":
    df_a = build_file_a()
    df_b = build_file_b()

    df_a.to_excel("data/raw/unique_employers.xlsx", index=False)
    df_b.to_excel("data/raw/lightcast_employers.xlsx", index=False)
    _format_raw("data/raw/unique_employers.xlsx")
    _format_raw("data/raw/lightcast_employers.xlsx")

    print(f"File A (unique_employers.xlsx): {len(df_a)} rows")
    print(f"File B (lightcast_employers.xlsx): {len(df_b)} rows")
